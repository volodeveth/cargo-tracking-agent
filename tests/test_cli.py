import json
import os
import subprocess
import sys
from pathlib import Path


def test_cli_reads_json_writes_json(tmp_path):
    inp = tmp_path / "in.json"
    inp.write_text(json.dumps({"shipments": [{"id": "a", "number": "080-38652331"}]}),
                   encoding="utf-8")
    out = tmp_path / "out.json"
    repo_root = Path(__file__).resolve().parents[1]
    env = os.environ.copy()
    env["PYTHONPATH"] = str(repo_root / "src")
    subprocess.run(
        [sys.executable, "-m", "tracking_agent.cli",
         "--input", str(inp), "--output", str(out)],
        check=True,
        cwd=str(repo_root),
        env=env,
    )
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["summary"]["total"] == 1
    assert data["results"][0]["input"]["number"] == "080-38652331"


def test_cli_reads_csv_prints_json(tmp_path):
    inp = tmp_path / "in.csv"
    inp.write_text("id,number\nb1,TLLU4912250\n", encoding="utf-8")
    repo_root = Path(__file__).resolve().parents[1]
    env = os.environ.copy()
    env["PYTHONPATH"] = str(repo_root / "src")
    env["PYTHONIOENCODING"] = "utf-8"
    result = subprocess.run(
        [sys.executable, "-m", "tracking_agent.cli", "--input", str(inp)],
        check=True,
        capture_output=True,
        cwd=str(repo_root),
        env=env,
    )
    data = json.loads(result.stdout.decode("utf-8"))
    assert data["summary"]["total"] == 1


def _env(repo_root):
    env = os.environ.copy()
    env["PYTHONPATH"] = str(repo_root / "src")
    env["PYTHONIOENCODING"] = "utf-8"
    env["LLM_ENABLED"] = "false"  # deterministic, no network in tests
    return env


def test_cli_xlsx_output(tmp_path):
    from openpyxl import load_workbook
    inp = tmp_path / "in.csv"
    inp.write_text("id,number\na1,080-38652331\n", encoding="utf-8")
    out = tmp_path / "out.xlsx"
    repo_root = Path(__file__).resolve().parents[1]
    subprocess.run(
        [sys.executable, "-m", "tracking_agent.cli",
         "--input", str(inp), "--output", str(out), "--format", "xlsx"],
        check=True, cwd=str(repo_root), env=_env(repo_root),
    )
    ws = load_workbook(str(out)).active
    assert ws["A1"].value == "id"
    assert ws["B2"].value == "080-38652331"


def test_cli_sheets_without_creds_falls_back_to_json(tmp_path):
    inp = tmp_path / "in.json"
    inp.write_text(json.dumps({"shipments": [{"id": "a", "number": "080-38652331"}]}),
                   encoding="utf-8")
    out = tmp_path / "out.json"
    repo_root = Path(__file__).resolve().parents[1]
    env = _env(repo_root)
    env["SHEETS_SPREADSHEET_ID"] = ""
    env["SHEETS_CREDENTIALS_PATH"] = ""
    result = subprocess.run(
        [sys.executable, "-m", "tracking_agent.cli",
         "--input", str(inp), "--output", str(out), "--format", "sheets"],
        check=True, capture_output=True, cwd=str(repo_root), env=env,
    )
    # no creds -> graceful fallback, still produces valid JSON and exits 0
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["summary"]["total"] == 1
    assert b"Sheets" in result.stderr or b"sheets" in result.stderr


def test_cli_invalid_number_produces_error(tmp_path):
    inp = tmp_path / "in.json"
    inp.write_text(json.dumps({"shipments": [{"id": "x", "number": "bad-number"}]}),
                   encoding="utf-8")
    out = tmp_path / "out.json"
    repo_root = Path(__file__).resolve().parents[1]
    env = os.environ.copy()
    env["PYTHONPATH"] = str(repo_root / "src")
    subprocess.run(
        [sys.executable, "-m", "tracking_agent.cli",
         "--input", str(inp), "--output", str(out)],
        check=True,
        cwd=str(repo_root),
        env=env,
    )
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["results"][0]["errors"][0]["code"] == "INVALID_FORMAT"
