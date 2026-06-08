from fastapi.testclient import TestClient
from tracking_agent.api.app import create_app


def test_track_endpoint():
    client = TestClient(create_app())
    resp = client.post("/track", json={"shipments": [
        {"id": "a", "number": "080-38652331"},
        {"id": "b", "number": "hello"}]})
    assert resp.status_code == 200
    data = resp.json()
    assert data["summary"]["total"] == 2
    assert data["results"][1]["errors"][0]["code"] == "INVALID_FORMAT"


def test_index_returns_html():
    client = TestClient(create_app())
    resp = client.get("/")
    assert resp.status_code == 200
    assert "text/html" in resp.headers["content-type"]
    assert "track/file" in resp.text


def test_track_file_csv():
    import io
    client = TestClient(create_app())
    csv_content = b"id,number\na1,080-38652331\n"
    resp = client.post(
        "/track/file",
        files={"file": ("batch.csv", io.BytesIO(csv_content), "text/csv")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["summary"]["total"] == 1


def test_track_file_csv_missing_number_column():
    """CSV without a 'number' column must return 422 with a helpful detail."""
    import io
    client = TestClient(create_app())
    csv_content = b"id,wrongcol\na1,foo\n"
    resp = client.post(
        "/track/file",
        files={"file": ("batch.csv", io.BytesIO(csv_content), "text/csv")},
    )
    assert resp.status_code == 422
    assert "number" in resp.json()["detail"]


def test_track_file_xlsx_missing_number_column(tmp_path):
    """XLSX without a 'number' column must return 422."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id"])
    ws.append(["a1"])
    path = tmp_path / "batch.xlsx"
    wb.save(str(path))
    client = TestClient(create_app())
    with open(str(path), "rb") as fh:
        resp = client.post(
            "/track/file",
            files={"file": ("batch.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 422
    assert "number" in resp.json()["detail"]


def test_track_file_unsupported_extension():
    """A file with an extension other than .csv/.xlsx must return 422."""
    import io
    client = TestClient(create_app())
    resp = client.post(
        "/track/file",
        files={"file": ("x.txt", io.BytesIO(b"some content"), "text/plain")},
    )
    assert resp.status_code == 422


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_track_endpoint_xlsx_format():
    import io
    from openpyxl import load_workbook
    client = TestClient(create_app())
    resp = client.post("/track?format=xlsx",
                       json={"shipments": [{"id": "a", "number": "080-38652331"}]})
    assert resp.status_code == 200
    assert _XLSX_MIME in resp.headers["content-type"]
    assert "attachment" in resp.headers.get("content-disposition", "")
    ws = load_workbook(io.BytesIO(resp.content)).active
    assert ws["A1"].value == "id"
    assert ws["B2"].value == "080-38652331"


def test_track_endpoint_defaults_to_json():
    client = TestClient(create_app())
    resp = client.post("/track", json={"shipments": [{"id": "a", "number": "080-38652331"}]})
    assert resp.status_code == 200
    assert "application/json" in resp.headers["content-type"]


def test_track_short_view():
    client = TestClient(create_app())
    resp = client.post("/track?view=short",
                       json={"shipments": [{"id": "a", "number": "080-38652331"}]})
    assert resp.status_code == 200
    data = resp.json()
    assert isinstance(data, list)
    assert data[0]["number"] == "080-38652331"
    assert "current_status" in data[0] and "last_event_at" in data[0]


def test_track_file_xlsx_export():
    import io
    from openpyxl import load_workbook
    client = TestClient(create_app())
    csv_content = b"id,number\na1,080-38652331\nb1,TLLU4912250\n"
    resp = client.post(
        "/track/file",
        data={"format": "xlsx"},
        files={"file": ("batch.csv", io.BytesIO(csv_content), "text/csv")},
    )
    assert resp.status_code == 200
    assert _XLSX_MIME in resp.headers["content-type"]
    ws = load_workbook(io.BytesIO(resp.content)).active
    assert ws["A1"].value == "id"
    assert ws.max_row == 3  # header + 2 rows
