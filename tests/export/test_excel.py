import io

from openpyxl import load_workbook
from tracking_agent.export.excel import export_results, export_results_bytes
from tracking_agent.models.schemas import ShipmentResult, ShipmentInput


def test_export_bytes_loads_as_workbook():
    results = [ShipmentResult(input=ShipmentInput(id="a", number="080-38652331"))]
    blob = export_results_bytes(results)
    assert isinstance(blob, bytes) and blob[:2] == b"PK"  # xlsx is a zip
    ws = load_workbook(io.BytesIO(blob)).active
    assert ws["A1"].value == "id"
    assert ws["B2"].value == "080-38652331"


def test_excel_has_header_and_rows(tmp_path):
    results = [ShipmentResult(input=ShipmentInput(id="a", number="080-38652331"))]
    path = tmp_path / "out.xlsx"
    export_results(results, str(path))
    wb = load_workbook(path)
    ws = wb.active
    assert ws["A1"].value == "id"
    assert ws["B2"].value == "080-38652331"


def test_excel_none_tracking_does_not_crash(tmp_path):
    results = [ShipmentResult(input=ShipmentInput(id="x", number="501-20285134"))]
    path = tmp_path / "none_tracking.xlsx"
    export_results(results, str(path))
    wb = load_workbook(path)
    ws = wb.active
    assert ws["A2"].value == "x"
    assert ws["E2"].value is None


def test_export_to_sheets_returns_false_without_creds():
    from tracking_agent.export.sheets import export_to_sheets
    assert export_to_sheets([], "", "") is False
