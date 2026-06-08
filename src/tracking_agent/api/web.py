from __future__ import annotations
import csv, io, uuid
from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse
from openpyxl import load_workbook
from ..models.schemas import ShipmentInput
from ..pipeline.builder import build_response

web_router = APIRouter()

_PAGE = """<!doctype html><html><body>
<h2>Cargo Tracking Agent</h2>
<form action="/track/file" method="post" enctype="multipart/form-data">
  <input type="file" name="file" accept=".csv,.xlsx"/>
  <button type="submit">Track</button>
</form></body></html>"""


class InvalidUploadError(ValueError):
    pass


@web_router.get("/", response_class=HTMLResponse)
async def index():
    return _PAGE


def _parse_csv(data: bytes) -> list[ShipmentInput]:
    reader = csv.DictReader(io.StringIO(data.decode("utf-8")))
    if reader.fieldnames is None or "number" not in reader.fieldnames:
        raise InvalidUploadError("Uploaded file must have a 'number' column")
    return [
        ShipmentInput(id=row.get("id"), number=row.get("number"))
        for row in reader
        if row.get("number")
    ]


def _parse_xlsx(data: bytes) -> list[ShipmentInput]:
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise InvalidUploadError("Uploaded file must have a 'number' column")
    headers = [(str(h).strip().lower() if h is not None else "") for h in rows[0]]
    if "number" not in headers:
        raise InvalidUploadError("Uploaded file must have a 'number' column")
    idx_id = headers.index("id") if "id" in headers else None
    idx_num = headers.index("number")
    out = []
    for r in rows[1:]:
        if r[idx_num] is None:
            continue
        out.append(ShipmentInput(id=str(r[idx_id]) if idx_id is not None else None,
                                 number=str(r[idx_num])))
    return out


@web_router.post("/track/file")
async def track_file(file: UploadFile = File(...)):
    filename = file.filename or ""
    if filename.endswith(".xlsx"):
        parse = _parse_xlsx
    elif filename.endswith(".csv"):
        parse = _parse_csv
    else:
        raise HTTPException(status_code=422, detail="Only .csv and .xlsx files are supported")
    data = await file.read()
    try:
        shipments = parse(data)
    except InvalidUploadError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    resp = await build_response(shipments, request_id=f"file-{uuid.uuid4().hex[:8]}")
    return JSONResponse(resp.model_dump(mode="json"))
